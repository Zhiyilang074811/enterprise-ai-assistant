from __future__ import annotations

import html
import os
import re
import csv
import json
import xml.etree.ElementTree as ET
from io import StringIO
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from bs4 import BeautifulSoup

from backend.langchain_components import (
    Document,
    load_documents_with_langchain,
    split_markdown_with_langchain,
)

try:
    import pdfplumber
except Exception:  # pragma: no cover
    pdfplumber = None

try:
    import pytesseract
    from PIL import Image
except Exception:  # pragma: no cover
    pytesseract = None
    Image = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

SUPPORTED_EXTENSIONS = {
    ".md",
    ".txt",
    ".html",
    ".htm",
    ".csv",
    ".json",
    ".xml",
    ".log",
    ".pdf",
    ".docx",
    ".pptx",
    ".xlsx",
    ".xls",
    ".png",
    ".jpg",
    ".jpeg",
}
TEXTLIKE_EXTENSIONS = {".md", ".txt", ".html", ".htm", ".csv", ".json", ".xml", ".log"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
PAGE_MERGE_MIN_CHARS = 280
PAGE_MERGE_MAX_CHARS = 2200
SEMANTIC_CHUNK_MIN_CHARS = 220
SEMANTIC_CHUNK_TARGET_CHARS = 720
SEMANTIC_CHUNK_MAX_CHARS = 1080
SEMANTIC_OVERLAP_MIN_CHARS = 40
SEMANTIC_OVERLAP_MAX_CHARS = 140
TABLE_BATCH_SIZE = 40

TIER_ALIAS = {
    "L1": "permanent",
    "L2": "seasonal",
    "L3": "hotfix",
    "permanent": "permanent",
    "seasonal": "seasonal",
    "hotfix": "hotfix",
}


@dataclass
class ParsedKnowledgeFile:
    filename: str
    canonical_tier: str
    original_suffix: str
    markdown_text: str
    source_type: str
    document_count: int
    chunk_count: int
    ingest_metadata: dict


@dataclass
class ParseOptions:
    parse_mode: str = "auto"
    pdf_mode: str = "auto"
    table_header_mode: str = "auto"


def normalize_parse_options(options: dict | None = None) -> ParseOptions:
    raw = options if isinstance(options, dict) else {}
    parse_mode = str(raw.get("parse_mode") or "auto").strip().lower()
    pdf_mode = str(raw.get("pdf_mode") or "auto").strip().lower()
    table_header_mode = str(raw.get("table_header_mode") or "auto").strip().lower()
    if parse_mode not in {"auto", "structure_first", "plain_text"}:
        parse_mode = "auto"
    if pdf_mode not in {"auto", "text", "scan", "table"}:
        pdf_mode = "auto"
    if table_header_mode not in {"auto", "first_row_header", "generate_header"}:
        table_header_mode = "auto"
    return ParseOptions(
        parse_mode=parse_mode,
        pdf_mode=pdf_mode,
        table_header_mode=table_header_mode,
    )


class UnsupportedDocumentError(ValueError):
    pass


def normalize_tier(value: str) -> str:
    tier = str(value or "").strip()
    return TIER_ALIAS.get(tier, tier)


def ensure_supported_extension(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise UnsupportedDocumentError(
            f"暂不支持 {suffix or '无后缀'} 文件，请上传 Markdown / TXT / HTML / CSV / PDF / DOCX / PPTX / XLSX / 图片"
        )
    return suffix


def _read_utf8(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except Exception:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _html_to_markdown_text(raw_html: str) -> str:
    soup = BeautifulSoup(raw_html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    lines: list[str] = []
    title = (soup.title.string or "").strip() if soup.title and soup.title.string else ""
    if title:
        lines.append(f"# {title}")
        lines.append("")
    for node in soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6", "p", "li", "tr"]):
        text = " ".join(node.get_text(" ", strip=True).split())
        if not text:
            continue
        name = node.name.lower()
        if name.startswith("h") and len(name) == 2 and name[1].isdigit():
            level = min(max(int(name[1]), 1), 6)
            lines.append(f"{'#' * level} {text}")
        elif name == "li":
            lines.append(f"- {text}")
        elif name == "tr":
            lines.append(f"| {text} |")
        else:
            lines.append(text)
        lines.append("")
    content = "\n".join(lines).strip()
    return html.unescape(content)


def _fallback_documents(path: Path, suffix: str) -> list[Document]:
    if suffix in {".md", ".txt"}:
        return [Document(page_content=_read_utf8(path), metadata={"source": str(path)})]
    if suffix in {".html", ".htm"}:
        return [Document(page_content=_html_to_markdown_text(_read_utf8(path)), metadata={"source": str(path)})]
    if suffix == ".csv":
        return _load_csv_documents(path)
    if suffix == ".json":
        return _load_json_documents(path)
    if suffix == ".xml":
        return _load_xml_documents(path)
    if suffix == ".log":
        return _load_log_documents(path)
    if suffix in IMAGE_EXTENSIONS:
        return _load_image_with_ocr(path)
    raise UnsupportedDocumentError(f"当前环境未安装 {suffix} 文档解析依赖，无法处理该文件")


def _load_plain_text_document(path: Path) -> list[Document]:
    return [Document(page_content=_read_utf8(path), metadata={"source": str(path), "structured": "plain_text"})]


def _sanitize_table_cell(value: object) -> str:
    text = str(value or "").replace("\r", " ").replace("\n", " ").strip()
    text = re.sub(r"\s{2,}", " ", text)
    return text


def _normalize_table_rows(rows: list[list[str]]) -> list[list[str]]:
    cleaned = [[_sanitize_table_cell(cell) for cell in row] for row in rows]
    cleaned = [row for row in cleaned if any(cell for cell in row)]
    max_len = max((len(row) for row in cleaned), default=0)
    if max_len <= 0:
        return []
    return [row + [""] * (max_len - len(row)) for row in cleaned]


def _apply_table_header_mode(rows: list[list[str]], header_mode: str) -> list[list[str]]:
    normalized = _normalize_table_rows(rows)
    if not normalized:
        return []
    if header_mode == "generate_header":
        col_count = len(normalized[0])
        return [[f"字段{i + 1}" for i in range(col_count)], *normalized]
    if header_mode == "first_row_header":
        return normalized
    return normalized


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    normalized = _normalize_table_rows(rows)
    if not normalized:
        return ""
    header = normalized[0]
    body = normalized[1:] if len(normalized) > 1 else []
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * len(header)) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(row[: len(header)]) + " |")
    return "\n".join(lines)


def _sample_rows(rows: list[list[str]], limit: int = 8) -> list[list[str]]:
    normalized = _normalize_table_rows(rows)
    if not normalized:
        return []
    return normalized[:limit]


def _load_csv_documents(path: Path, header_mode: str = "auto") -> list[Document]:
    raw = _read_utf8(path)
    try:
        sample = raw[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(StringIO(raw), dialect=dialect)
    rows = _apply_table_header_mode([row for row in reader], header_mode)
    if not rows:
        return [Document(page_content="空表格", metadata={"source": str(path), "structured": "table", "table_kind": "csv"})]
    documents: list[Document] = []
    header = rows[0]
    body = rows[1:] or [[]]
    for offset in range(0, len(body), TABLE_BATCH_SIZE):
        batch = body[offset : offset + TABLE_BATCH_SIZE]
        block_rows = [header, *batch]
        documents.append(
            Document(
                page_content=_rows_to_markdown_table(block_rows),
                metadata={
                    "source": str(path),
                    "structured": "table",
                    "table_kind": "csv",
                    "row_start": offset + 1,
                    "row_end": offset + len(batch),
                },
            )
        )
    return documents


def _json_value_to_lines(value: object, prefix: str = "") -> list[str]:
    lines: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            clean_key = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(item, (dict, list)):
                lines.append(f"## {clean_key}")
                lines.extend(_json_value_to_lines(item, clean_key))
            else:
                lines.append(f"- {clean_key}: {_sanitize_table_cell(item)}")
        return lines
    if isinstance(value, list):
        for index, item in enumerate(value):
            clean_key = f"{prefix}[{index}]"
            if isinstance(item, (dict, list)):
                lines.append(f"## {clean_key}")
                lines.extend(_json_value_to_lines(item, clean_key))
            else:
                lines.append(f"- {clean_key}: {_sanitize_table_cell(item)}")
        return lines
    if prefix:
        return [f"- {prefix}: {_sanitize_table_cell(value)}"]
    return [f"- value: {_sanitize_table_cell(value)}"]


def _load_json_documents(path: Path) -> list[Document]:
    raw = _read_utf8(path)
    try:
        data = json.loads(raw)
    except Exception:
        return [Document(page_content=raw, metadata={"source": str(path), "structured": "json_raw"})]
    if isinstance(data, list) and data and all(isinstance(item, dict) for item in data[:50]):
        header = sorted({str(key) for item in data[:50] for key in item.keys()})
        rows = [header]
        for item in data[:200]:
            rows.append([_sanitize_table_cell(item.get(col, "")) for col in header])
        table_text = _rows_to_markdown_table(rows)
        return [Document(page_content=table_text, metadata={"source": str(path), "structured": "table", "table_kind": "json"})]
    lines = _json_value_to_lines(data)
    return [Document(page_content="\n".join(lines).strip(), metadata={"source": str(path), "structured": "json"})]


def _xml_element_to_lines(element: ET.Element, level: int = 1) -> list[str]:
    tag = re.sub(r"^\{.*\}", "", element.tag)
    heading = f"{'#' * min(level, 6)} <{tag}>"
    lines = [heading]
    attrs = " ".join(f'{key}="{_sanitize_table_cell(value)}"' for key, value in element.attrib.items())
    if attrs:
        lines.append(f"- attributes: {attrs}")
    text = _sanitize_table_cell(element.text or "")
    if text:
        lines.append(f"- text: {text}")
    children = list(element)
    for child in children[:60]:
        lines.extend(_xml_element_to_lines(child, level + 1))
    return lines


def _load_xml_documents(path: Path) -> list[Document]:
    raw = _read_utf8(path)
    try:
        root = ET.fromstring(raw)
    except Exception:
        return [Document(page_content=raw, metadata={"source": str(path), "structured": "xml_raw"})]
    lines = _xml_element_to_lines(root)
    return [Document(page_content="\n".join(lines).strip(), metadata={"source": str(path), "structured": "xml"})]


def _load_log_documents(path: Path) -> list[Document]:
    raw = _read_utf8(path)
    lines = [line.rstrip() for line in raw.replace("\r", "").splitlines() if line.strip()]
    if not lines:
        return [Document(page_content="空日志", metadata={"source": str(path), "structured": "log"})]
    documents: list[Document] = []
    batch_size = 80
    for offset in range(0, len(lines), batch_size):
        batch = lines[offset : offset + batch_size]
        documents.append(
            Document(
                page_content="\n".join(batch),
                metadata={
                    "source": str(path),
                    "structured": "log",
                    "line_start": offset + 1,
                    "line_end": offset + len(batch),
                },
            )
        )
    return documents


def _load_excel_documents(path: Path, header_mode: str = "auto") -> list[Document]:
    if load_workbook is None or path.suffix.lower() == ".xls":
        docs = load_documents_with_langchain(path)
        if docs:
            return [Document(page_content=_clean_text(doc.page_content), metadata=dict(doc.metadata or {})) for doc in docs if _clean_text(doc.page_content)]
        return _fallback_documents(path, path.suffix.lower())
    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        docs = load_documents_with_langchain(path)
        if docs:
            return [Document(page_content=_clean_text(doc.page_content), metadata=dict(doc.metadata or {})) for doc in docs if _clean_text(doc.page_content)]
        return _fallback_documents(path, path.suffix.lower())

    documents: list[Document] = []
    for sheet in workbook.worksheets:
        raw_rows = []
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append([_sanitize_table_cell(cell) for cell in row])
        rows = _apply_table_header_mode(raw_rows, header_mode)
        if not rows:
            continue
        header = rows[0]
        body = rows[1:] or [[]]
        for offset in range(0, len(body), TABLE_BATCH_SIZE):
            batch = body[offset : offset + TABLE_BATCH_SIZE]
            block_rows = [header, *batch]
            documents.append(
                Document(
                    page_content=f"### 工作表：{sheet.title}\n\n{_rows_to_markdown_table(block_rows)}",
                    metadata={
                        "source": str(path),
                        "structured": "table",
                        "table_kind": "excel",
                        "sheet": sheet.title,
                        "row_start": offset + 1,
                        "row_end": offset + len(batch),
                    },
                )
            )
    workbook.close()
    return documents or [Document(page_content="空工作簿", metadata={"source": str(path), "structured": "table", "table_kind": "excel"})]


def _load_image_with_ocr(path: Path) -> list[Document]:
    """图片文件走 OCR，方便企业上传扫描件和截图。"""
    if pytesseract is None or Image is None:
        raise UnsupportedDocumentError("当前环境未安装 OCR 依赖，暂不支持图片识别，请先上传文字版或安装 Tesseract / Pillow")
    text = pytesseract.image_to_string(Image.open(path), lang="chi_sim+eng")
    cleaned = _clean_text(text)
    return [Document(page_content=cleaned or "图片未识别出可用文字", metadata={"source": str(path), "ocr": True})]


def _ocr_pdf_with_pdfplumber(path: Path) -> list[Document]:
    """针对扫描版 PDF 做兜底 OCR，优先保证有字可检索。"""
    if pdfplumber is None or pytesseract is None:
        return []
    documents: list[Document] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                page_text = _clean_text(page.extract_text() or "")
                if page_text:
                    documents.append(
                        Document(
                            page_content=page_text,
                            metadata={"source": str(path), "page": page_index - 1, "ocr": False},
                        )
                    )
                    continue
                try:
                    rendered = page.to_image(resolution=220)
                    image = rendered.original
                except Exception:
                    continue
                ocr_text = _clean_text(pytesseract.image_to_string(image, lang="chi_sim+eng"))
                if not ocr_text:
                    continue
                documents.append(
                    Document(
                        page_content=ocr_text,
                        metadata={"source": str(path), "page": page_index - 1, "ocr": True},
                    )
                )
    except Exception:
        return []
    return documents


def _extract_pdf_tables(path: Path) -> list[str]:
    """提取 PDF 表格，尽量把企业制度、报价单、排班表这类结构化内容保留下来。"""
    if pdfplumber is None:
        return []
    table_blocks: list[str] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                for table_index, table in enumerate(page.extract_tables() or [], start=1):
                    rows = []
                    for row in table or []:
                        cleaned = [re.sub(r"\s+", " ", str(cell or "").strip()) for cell in row]
                        if any(cleaned):
                            rows.append(cleaned)
                    if not rows:
                        continue
                    header = rows[0]
                    body = rows[1:] if len(rows) > 1 else []
                    lines = [f"## 第 {page_index} 页表格 {table_index}", ""]
                    lines.append("| " + " | ".join(header) + " |")
                    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                    for row in body:
                        padded = row + [""] * (len(header) - len(row))
                        lines.append("| " + " | ".join(padded[: len(header)]) + " |")
                    table_blocks.append("\n".join(lines).strip())
    except Exception:
        return []
    return table_blocks


def _merge_cross_page_documents(docs: list[Document]) -> list[Document]:
    """
    合并跨页断开的短文本，减少企业 PDF / Word / PPT 在分页处被切碎的情况。
    规则尽量保守：只合并短页，避免把整章文档揉成一块。
    """
    if len(docs) <= 1:
        return docs
    merged: list[Document] = []
    buffer_doc: Document | None = None
    for doc in docs:
        content = _clean_text(doc.page_content)
        if not content:
            continue
        metadata = dict(doc.metadata or {})
        current = Document(page_content=content, metadata=metadata)
        if buffer_doc is None:
            buffer_doc = current
            continue
        buffer_text = _clean_text(buffer_doc.page_content)
        same_source = buffer_doc.metadata.get("source") == current.metadata.get("source")
        page_a = buffer_doc.metadata.get("page")
        page_b = current.metadata.get("page")
        consecutive_pages = isinstance(page_a, int) and isinstance(page_b, int) and page_b == page_a + 1
        should_merge = (
            same_source
            and consecutive_pages
            and len(buffer_text) < PAGE_MERGE_MIN_CHARS
            and len(buffer_text) + len(content) <= PAGE_MERGE_MAX_CHARS
        )
        if should_merge:
            buffer_doc = Document(
                page_content=f"{buffer_text}\n\n{content}",
                metadata={**buffer_doc.metadata, "page_end": page_b, "merged_pages": True},
            )
            continue
        merged.append(buffer_doc)
        buffer_doc = current
    if buffer_doc is not None:
        merged.append(buffer_doc)
    return merged


def _load_pdf_documents(path: Path, options: ParseOptions | None = None) -> list[Document]:
    """PDF 优先走文本解析，文本过少时再用结构化解析或 OCR 兜底。"""
    opts = options or ParseOptions()
    candidates = load_documents_with_langchain(path)
    effective_pdf_mode = opts.pdf_mode
    if opts.parse_mode == "plain_text" and effective_pdf_mode == "auto":
        effective_pdf_mode = "text"
    if effective_pdf_mode == "scan":
        ocr_docs = _ocr_pdf_with_pdfplumber(path)
        if ocr_docs:
            return _merge_cross_page_documents(ocr_docs)
        return _merge_cross_page_documents(candidates)
    if effective_pdf_mode == "table":
        table_blocks = _extract_pdf_tables(path)
        if table_blocks:
            return [
                Document(
                    page_content=block,
                    metadata={"source": str(path), "structured": "table", "table_kind": "pdf"},
                )
                for block in table_blocks
            ]
    total_chars = sum(len(_clean_text(doc.page_content)) for doc in candidates)
    if effective_pdf_mode == "text" and candidates:
        return _merge_cross_page_documents(candidates)
    if total_chars >= 180:
        return _merge_cross_page_documents(candidates)
    ocr_docs = _ocr_pdf_with_pdfplumber(path)
    if ocr_docs:
        return _merge_cross_page_documents(ocr_docs)
    return _merge_cross_page_documents(candidates)


def _load_word_documents(path: Path) -> list[Document]:
    """Word 文档优先保结构，解析失败再回退到简单文本提取。"""
    docs = load_documents_with_langchain(path)
    if docs:
        return [Document(page_content=_clean_text(doc.page_content), metadata=dict(doc.metadata or {})) for doc in docs if _clean_text(doc.page_content)]
    return _fallback_documents(path, path.suffix.lower())


def _load_html_documents(path: Path) -> list[Document]:
    docs = load_documents_with_langchain(path)
    if docs:
        return [Document(page_content=_clean_text(doc.page_content), metadata=dict(doc.metadata or {})) for doc in docs if _clean_text(doc.page_content)]
    return _fallback_documents(path, path.suffix.lower())


def load_documents(path: Path, options: ParseOptions | None = None) -> list[Document]:
    opts = options or ParseOptions()
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _load_pdf_documents(path, opts)
    if suffix == ".docx":
        return _load_word_documents(path)
    if opts.parse_mode == "plain_text" and suffix in {".csv", ".json", ".xml", ".log"}:
        return _load_plain_text_document(path)
    if suffix == ".csv":
        return _load_csv_documents(path, opts.table_header_mode)
    if suffix in {".xlsx", ".xls"}:
        return _load_excel_documents(path, opts.table_header_mode)
    if suffix == ".json":
        return _load_json_documents(path)
    if suffix == ".xml":
        return _load_xml_documents(path)
    if suffix == ".log":
        return _load_log_documents(path)
    if suffix in {".pptx", ".md", ".txt"}:
        docs = load_documents_with_langchain(path)
        if docs:
            return [Document(page_content=_clean_text(doc.page_content), metadata=dict(doc.metadata or {})) for doc in docs if _clean_text(doc.page_content)]
    if suffix in {".html", ".htm"}:
        return _load_html_documents(path)
    return _fallback_documents(path, suffix)


def _clean_text(text: str) -> str:
    text = text.replace("\r", "")
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"([^\n])-\n([^\n])", r"\1\2", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _normalize_markdown_structure(markdown_text: str) -> str:
    """
    统一标题和正文之间的空行，方便后续做标题分块和章节分块。
    """
    lines = markdown_text.replace("\r", "").splitlines()
    normalized: list[str] = []
    previous_blank = False
    for raw_line in lines:
        line = raw_line.rstrip()
        stripped = line.strip()
        if stripped.startswith("#"):
            if normalized and normalized[-1] != "":
                normalized.append("")
            normalized.append(stripped)
            previous_blank = False
            continue
        if not stripped:
            if not previous_blank:
                normalized.append("")
            previous_blank = True
            continue
        normalized.append(line)
        previous_blank = False
    text = "\n".join(normalized).strip()
    return re.sub(r"\n{3,}", "\n\n", text)


def _documents_to_markdown(docs: Iterable[Document], filename: str, suffix: str) -> str:
    docs = list(docs)
    if suffix == ".md":
        merged = "\n\n".join(_clean_text(doc.page_content) for doc in docs if _clean_text(doc.page_content))
        return merged

    title = Path(filename).stem
    lines = [f"# {title}", "", f"来源文件：{filename}", ""]
    for index, doc in enumerate(docs, start=1):
        content = _clean_text(doc.page_content)
        if not content:
            continue
        page_num = doc.metadata.get("page") if isinstance(doc.metadata, dict) else None
        if page_num is not None:
            lines.append(f"## 第 {int(page_num) + 1} 页")
        elif len(docs) > 1:
            lines.append(f"## 第 {index} 段")
        lines.append(content)
        lines.append("")
    return "\n".join(lines).strip()


def _append_structured_extras(markdown_text: str, *, path: Path, suffix: str) -> str:
    """把 OCR / 表格等结构化附加内容并到 Markdown，保证后续切块时能一起参与检索。"""
    blocks = [markdown_text.strip()]
    if suffix == ".pdf":
        table_blocks = _extract_pdf_tables(path)
        if table_blocks:
            blocks.append("## 附加结构化表格")
            blocks.append("")
            blocks.extend(table_blocks)
    return "\n\n".join(block for block in blocks if block).strip()


def _extract_outline(markdown_text: str, limit: int = 12) -> list[str]:
    items: list[str] = []
    for line in markdown_text.splitlines():
        stripped = line.strip()
        if not stripped.startswith("#"):
            continue
        heading = re.sub(r"^#+\s*", "", stripped).strip()
        if heading:
            items.append(heading)
        if len(items) >= limit:
            break
    return items


def _collect_markdown_tables(markdown_text: str, limit: int = 3) -> list[dict]:
    tables: list[dict] = []
    current_rows: list[list[str]] = []
    for line in markdown_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            current_rows.append([cell.strip() for cell in stripped.strip("|").split("|")])
            continue
        if current_rows:
            if len(current_rows) >= 2:
                sample = _sample_rows(current_rows[:10], limit=6)
                tables.append(
                    {
                        "columns": sample[0] if sample else [],
                        "rows": sample[1:] if len(sample) > 1 else [],
                        "row_count": max(len(current_rows) - 2, 0),
                    }
                )
                if len(tables) >= limit:
                    return tables
            current_rows = []
    if current_rows and len(current_rows) >= 2 and len(tables) < limit:
        sample = _sample_rows(current_rows[:10], limit=6)
        tables.append(
            {
                "columns": sample[0] if sample else [],
                "rows": sample[1:] if len(sample) > 1 else [],
                "row_count": max(len(current_rows) - 2, 0),
            }
        )
    return tables


def _build_tabular_preview(path: Path, suffix: str, header_mode: str = "auto") -> dict:
    if suffix == ".csv":
        raw = _read_utf8(path)
        try:
            sample = raw[:4096]
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except Exception:
            dialect = csv.excel
        rows = _apply_table_header_mode(list(csv.reader(StringIO(raw), dialect=dialect)), header_mode)
        sample_rows = _sample_rows(rows, limit=7)
        return {
            "display_mode": "table",
            "table_kind": "csv",
            "sheet_count": 1,
            "tables": [
                {
                    "name": Path(path).name,
                    "columns": sample_rows[0] if sample_rows else [],
                    "rows": sample_rows[1:] if len(sample_rows) > 1 else [],
                    "row_count": max(len(rows) - 1, 0),
                }
            ],
        }
    if suffix in {".xlsx", ".xls"} and load_workbook is not None and suffix != ".xls":
        try:
            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception:
            workbook = None
        if workbook is not None:
            tables = []
            sheet_count = len(workbook.sheetnames)
            for sheet in workbook.worksheets:
                rows = _apply_table_header_mode(
                    [[_sanitize_table_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
                    , header_mode
                )
                if not rows:
                    continue
                sample_rows = _sample_rows(rows, limit=7)
                tables.append(
                    {
                        "name": sheet.title,
                        "columns": sample_rows[0] if sample_rows else [],
                        "rows": sample_rows[1:] if len(sample_rows) > 1 else [],
                        "row_count": max(len(rows) - 1, 0),
                    }
                )
                if len(tables) >= 4:
                    break
            workbook.close()
            if tables:
                return {
                    "display_mode": "table",
                    "table_kind": "excel",
                    "sheet_count": sheet_count,
                    "tables": tables,
                }
    return {
        "display_mode": "table",
        "table_kind": "spreadsheet",
        "sheet_count": 0,
        "tables": _collect_markdown_tables(_read_utf8(path) if suffix in TEXTLIKE_EXTENSIONS else ""),
    }


def _build_ingest_metadata(
    *,
    filename: str,
    suffix: str,
    markdown_text: str,
    path: Path,
    documents: list[Document],
    options: ParseOptions,
) -> dict:
    effective_pdf_mode = options.pdf_mode
    if suffix == ".pdf" and options.parse_mode == "plain_text" and effective_pdf_mode == "auto":
        effective_pdf_mode = "text"
    parser_chain = ["format_router", "langchain_loader", "markdown_normalizer", "semantic_chunker"]
    if suffix == ".pdf":
        if effective_pdf_mode == "scan":
            parser_chain.insert(1, "pdf_scan_ocr")
        elif effective_pdf_mode == "table":
            parser_chain.insert(1, "pdf_table_first")
        elif effective_pdf_mode == "text":
            parser_chain.insert(1, "pdf_text_first")
        else:
            parser_chain.insert(2, "pdf_table_extractor")
        display_mode = "pdf"
        preview = {
            "display_mode": "pdf",
            "page_count": len(documents),
            "outline": _extract_outline(markdown_text),
            "tables": _collect_markdown_tables(markdown_text, limit=3),
        }
    elif suffix in {".csv", ".xlsx", ".xls"}:
        parser_chain.insert(1, "tabular_structured_parser")
        display_mode = "table"
        preview = _build_tabular_preview(path, suffix, options.table_header_mode)
    elif suffix in IMAGE_EXTENSIONS:
        parser_chain.insert(1, "ocr")
        display_mode = "image_ocr"
        preview = {
            "display_mode": "image_ocr",
            "ocr_enabled": True,
            "outline": _extract_outline(markdown_text, limit=6),
            "excerpt": _clean_text(markdown_text)[:320],
        }
    elif suffix in {".docx", ".pptx"}:
        display_mode = "office"
        preview = {
            "display_mode": "office",
            "outline": _extract_outline(markdown_text),
            "tables": _collect_markdown_tables(markdown_text, limit=2),
        }
    elif suffix in {".html", ".htm"}:
        display_mode = "web"
        preview = {
            "display_mode": "web",
            "outline": _extract_outline(markdown_text),
            "excerpt": _clean_text(markdown_text)[:320],
        }
    elif suffix == ".json":
        display_mode = "json"
        preview = {
            "display_mode": "json",
            "outline": _extract_outline(markdown_text),
            "excerpt": _clean_text(markdown_text)[:320],
            "tables": _collect_markdown_tables(markdown_text, limit=2),
        }
    elif suffix == ".xml":
        display_mode = "xml"
        preview = {
            "display_mode": "xml",
            "outline": _extract_outline(markdown_text),
            "excerpt": _clean_text(markdown_text)[:320],
        }
    elif suffix == ".log":
        display_mode = "log"
        preview = {
            "display_mode": "log",
            "excerpt": _clean_text(markdown_text)[:320],
        }
    else:
        display_mode = "text"
        preview = {
            "display_mode": "text",
            "outline": _extract_outline(markdown_text),
            "excerpt": _clean_text(markdown_text)[:320],
        }
    return {
        "source_name": filename,
        "source_type": suffix.lstrip("."),
        "original_suffix": suffix,
        "display_mode": display_mode,
        "parser_chain": parser_chain,
        "parse_mode": options.parse_mode,
        "pdf_mode": effective_pdf_mode,
        "table_header_mode": options.table_header_mode,
        "preview": preview,
    }


def _protect_tables(markdown_text: str, metadata: dict | None = None) -> list[Document]:
    """优先把 Markdown 表格视为独立知识块，避免后续递归切分把表格切碎。"""
    segments: list[Document] = []
    current_lines: list[str] = []
    current_is_table = False
    base_metadata = dict(metadata or {})

    def flush() -> None:
        nonlocal current_lines, current_is_table
        content = "\n".join(current_lines).strip()
        if not content:
            current_lines = []
            current_is_table = False
            return
        doc_metadata = dict(base_metadata)
        if current_is_table:
            doc_metadata["structured"] = "table"
        segments.append(Document(page_content=content, metadata=doc_metadata))
        current_lines = []
        current_is_table = False

    for line in markdown_text.splitlines():
        stripped = line.strip()
        is_table_line = stripped.startswith("|") and stripped.endswith("|")
        if current_lines and is_table_line != current_is_table:
            flush()
        current_is_table = is_table_line
        current_lines.append(line)
    flush()
    return [doc for doc in segments if doc.page_content.strip()]


def _split_markdown_sections(markdown_text: str) -> list[Document]:
    """
    先按标题和章节切大块，并保留标题路径，方便后续语义分块时保住上下文。
    """
    normalized = _normalize_markdown_structure(markdown_text)
    sections: list[Document] = []
    heading_stack: list[str] = []
    current_heading = ""
    current_lines: list[str] = []

    def flush() -> None:
        nonlocal current_lines, current_heading
        content = "\n".join(current_lines).strip()
        if not content:
            current_lines = []
            return
        heading_path = " > ".join(part for part in heading_stack if part)
        metadata = {
            "heading": current_heading,
            "heading_path": heading_path,
        }
        sections.extend(_protect_tables(content, metadata))
        current_lines = []

    for line in normalized.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            match = re.match(r"^(#{1,6})\s+(.*)$", stripped)
            if match:
                flush()
                level = len(match.group(1))
                heading_text = match.group(2).strip()
                while len(heading_stack) >= level:
                    heading_stack.pop()
                heading_stack.append(heading_text)
                current_heading = heading_text
                current_lines = [stripped]
                continue
        current_lines.append(line)
    flush()
    return sections or [Document(page_content=normalized, metadata={})]


def _split_sentences(text: str) -> list[str]:
    parts = re.split(r"(?<=[。！？!?；;])(?:\s+|(?=[^\s]))|\n{2,}", text)
    sentences = [part.strip() for part in parts if part and part.strip()]
    if sentences:
        return sentences
    return [text.strip()] if text.strip() else []


def _tokenize_for_similarity(text: str) -> set[str]:
    tokens = re.findall(r"[\u4e00-\u9fff]{1,4}|[A-Za-z0-9_]{2,}", text.lower())
    return set(tokens)


def _estimate_semantic_overlap_chars(content: str, sentences: list[str], metadata: dict | None = None) -> int:
    text_len = len(content)
    sentence_count = len(sentences)
    heading_path = str((metadata or {}).get("heading_path") or "").strip()
    heading_depth = heading_path.count(" > ") + 1 if heading_path else 0
    base = 60
    if text_len >= 1800:
        base += 20
    if text_len >= 3200:
        base += 20
    if sentence_count >= 10:
        base += 10
    if sentence_count >= 18:
        base += 10
    if heading_depth <= 1:
        base += 10
    if re.search(r"[：:；;]", content):
        base += 10
    return max(SEMANTIC_OVERLAP_MIN_CHARS, min(SEMANTIC_OVERLAP_MAX_CHARS, base))


def _build_overlap_tail(sentences: list[str], overlap_chars: int) -> tuple[list[str], set[str], int]:
    if not sentences or overlap_chars <= 0:
        return [], set(), 0
    selected: list[str] = []
    total = 0
    for sentence in reversed(sentences):
        sentence = str(sentence or "").strip()
        if not sentence:
            continue
        selected.insert(0, sentence)
        total += len(sentence)
        if total >= overlap_chars:
            break
    merged = "\n".join(selected).strip()
    return selected, _tokenize_for_similarity(merged), len(merged)


def _semantic_split_document(doc: Document) -> list[Document]:
    """
    用轻量语义策略把章节块再切小：
    - 优先按句子聚合
    - 结合关键词重叠度判断是否继续并块
    - 避免切得过碎
    """
    content = _clean_text(doc.page_content)
    if len(content) <= SEMANTIC_CHUNK_MAX_CHARS:
        return [Document(page_content=content, metadata=dict(doc.metadata or {}))]

    sentences = _split_sentences(content)
    if not sentences:
        return []

    chunks: list[Document] = []
    metadata = dict(doc.metadata or {})
    overlap_chars = _estimate_semantic_overlap_chars(content, sentences, metadata)
    current_sentences: list[str] = []
    current_tokens: set[str] = set()
    current_len = 0

    def flush() -> None:
        nonlocal current_sentences, current_tokens, current_len
        chunk_text = "\n".join(current_sentences).strip()
        if chunk_text:
            chunk_meta = dict(metadata)
            chunk_meta["chunk_overlap_chars"] = overlap_chars
            chunks.append(Document(page_content=chunk_text, metadata=chunk_meta))
        overlap_tail, overlap_tokens, overlap_len = _build_overlap_tail(current_sentences, overlap_chars)
        current_sentences = overlap_tail
        current_tokens = overlap_tokens
        current_len = overlap_len

    for sentence in sentences:
        sentence_tokens = _tokenize_for_similarity(sentence)
        sentence_len = len(sentence)
        overlap = len(current_tokens & sentence_tokens) if current_tokens else 0
        should_start_new = False

        if current_sentences:
            if current_len >= SEMANTIC_CHUNK_MAX_CHARS:
                should_start_new = True
            elif current_len >= SEMANTIC_CHUNK_TARGET_CHARS and overlap == 0:
                should_start_new = True
            elif current_len >= SEMANTIC_CHUNK_MIN_CHARS and sentence.startswith("#"):
                should_start_new = True

        if should_start_new:
            flush()

        current_sentences.append(sentence)
        current_tokens |= sentence_tokens
        current_len += sentence_len

    chunk_text = "\n".join(current_sentences).strip()
    if chunk_text:
        chunk_meta = dict(metadata)
        chunk_meta["chunk_overlap_chars"] = overlap_chars
        chunks.append(Document(page_content=chunk_text, metadata=chunk_meta))
    return chunks


def _apply_semantic_chunking(docs: list[Document]) -> list[Document]:
    final_docs: list[Document] = []
    for doc in docs:
        if doc.metadata.get("structured") == "table":
            final_docs.append(doc)
            continue
        final_docs.extend(_semantic_split_document(doc))
    return final_docs


def split_documents_for_stats(markdown_text: str) -> list[Document]:
    """企业文档友好的分块策略：标题分块 -> 章节分块 -> 表格保护 -> 语义分块。"""
    normalized = _normalize_markdown_structure(markdown_text)
    docs = split_markdown_with_langchain(normalized)
    docs = [
        Document(
            page_content=_clean_text(doc.page_content),
            metadata={key: value for key, value in dict(doc.metadata or {}).items() if value},
        )
        for doc in docs
        if _clean_text(doc.page_content)
    ] or _split_markdown_sections(normalized)

    section_docs: list[Document] = []
    for doc in docs:
        if doc.metadata.get("structured") == "table":
            section_docs.append(doc)
            continue
        section_docs.extend(_split_markdown_sections(doc.page_content))
    docs = section_docs or docs

    semantic_docs = _apply_semantic_chunking(docs)
    return semantic_docs or docs


def parse_uploaded_knowledge_file(
    *,
    filename: str,
    raw_bytes: bytes,
    tier: str,
    temp_dir: Path,
    parse_options: dict | None = None,
) -> ParsedKnowledgeFile:
    suffix = ensure_supported_extension(filename)
    canonical_tier = normalize_tier(tier)
    options = normalize_parse_options(parse_options)
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_path = temp_dir / filename
    temp_path.write_bytes(raw_bytes)
    documents = load_documents(temp_path, options)
    markdown_text = _documents_to_markdown(documents, filename=filename, suffix=suffix)
    markdown_text = _append_structured_extras(markdown_text, path=temp_path, suffix=suffix)
    chunk_docs = split_documents_for_stats(markdown_text)
    output_name = f"{Path(filename).stem}.md"
    ingest_metadata = _build_ingest_metadata(
        filename=filename,
        suffix=suffix,
        markdown_text=markdown_text,
        path=temp_path,
        documents=documents,
        options=options,
    )
    return ParsedKnowledgeFile(
        filename=output_name,
        canonical_tier=canonical_tier,
        original_suffix=suffix,
        markdown_text=markdown_text,
        source_type=suffix.lstrip("."),
        document_count=len(documents),
        chunk_count=len(chunk_docs),
        ingest_metadata=ingest_metadata,
    )
